from fastapi import FastAPI, UploadFile, File
from fastapi.responses import PlainTextResponse
import json
from openpyxl import load_workbook
from datetime import datetime
import requests
import tempfile

app = FastAPI()

# ==== OpenRouter 配置 ====
OPENROUTER_API_KEY = "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"
OPENROUTER_API_URL = "https://openrouter.ai/api/v1/chat/completions"
OPENROUTER_MODEL = "openai/gpt-4o"
MAX_PROMPT_LENGTH = 12000


def query_openrouter(prompt: str, model=OPENROUTER_MODEL):
    headers = {
        "Authorization": f"Bearer {OPENROUTER_API_KEY}",
        "Content-Type": "application/json"
    }
    payload = {
        "model": model,
        "messages": [
            {"role": "system", "content": "你是一位数据分析师，请将用户提供的 JSON 数据逐项转述为可读性强、细致完整的中文描述，确保不遗漏任何信息。"},
            {"role": "user", "content": prompt}
        ]
    }
    response = requests.post(OPENROUTER_API_URL, headers=headers, json=payload)
    try:
        return response.json()['choices'][0]['message']['content']
    except Exception as e:
        raise RuntimeError(f"‼️ OpenRouter 响应异常：{response.text}") from e


@app.post("/trans_content", response_class=PlainTextResponse)
async def trans_content(file: UploadFile = File(...)):
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(await file.read())
        tmp_path = tmp.name

    workbook = load_workbook(filename=tmp_path, data_only=True)
    all_json_results = {}

    # === Period Sheet 处理 ===
    def store_period_section(section, rows):
        if section == "[Roster Period]":
            section_data = {
                str(row[0]).strip(): row[1].strftime('%Y-%m-%d %H:%M:%S') if isinstance(row[1], datetime) else str(row[1]).strip()
                for row in rows if row[0] and row[1]
            }
            period_result[section] = section_data
        elif section == "[Monthly pattern]":
            monthly_data = {"Date": "DOW"}
            for row in rows:
                if isinstance(row[0], datetime) and row[1]:
                    monthly_data[row[0].strftime('%Y-%m-%d %H:%M:%S')] = row[1]
            period_result[section] = monthly_data
        elif section == "[Weekly pattern]":
            if not rows:
                period_result[section] = {}
                return
            headers = [str(c).strip() if c else "" for c in rows[0]]
            weekly_data = {}
            for row in rows[1:]:
                if not row[0]:
                    continue
                weekly_data[str(row[0]).strip()] = [str(cell).strip() for cell in row[1:] if cell]
            period_result[section] = weekly_data
        elif section == "[Day Tagging]":
            period_result[section] = {}

    period_result = {}
    sheet = workbook["Period"]
    current_section, data = None, []

    for row in sheet.iter_rows(values_only=True):
        if all(cell is None for cell in row):
            continue
        first_cell = str(row[0]).strip() if row[0] else ""
        if first_cell.startswith("[") and first_cell.endswith("]"):
            if current_section:
                store_period_section(current_section, data)
            current_section = first_cell
            data = []
        elif current_section:
            data.append(row)
    if current_section:
        store_period_section(current_section, data)
    all_json_results.update(period_result)

    # === Staff Sheet ===
    sheet = workbook["Staff"]
    profile_cols = list(range(2, 8))
    manpower_cols = list(range(8, 19))
    header_row = 4
    data_start_row = 5
    data_end_row = 24
    profile_headers = [sheet.cell(row=header_row, column=col).value for col in profile_cols]
    manpower_headers = [sheet.cell(row=header_row, column=col).value for col in manpower_cols]
    staff_profile_data, staff_manpower_data = [], []

    for row in range(data_start_row, data_end_row + 1):
        profile_row = {profile_headers[i]: sheet.cell(row=row, column=col).value for i, col in enumerate(profile_cols)}
        manpower_row = {manpower_headers[i]: sheet.cell(row=row, column=col).value for i, col in enumerate(manpower_cols)}
        staff_profile_data.append(profile_row)
        staff_manpower_data.append(manpower_row)

    all_json_results["[Staff Profile]"] = staff_profile_data
    all_json_results["[Staff Manpower]"] = staff_manpower_data

    # === Duty Sheet ===
    sheet = workbook["Duty"]
    duty_profile_cols = list(range(2, 11))
    staff_required_cols = list(range(12, 21))
    duty_profile_headers = [sheet.cell(row=header_row, column=col).value for col in duty_profile_cols]
    staff_required_headers = [sheet.cell(row=header_row, column=col).value for col in staff_required_cols]
    duty_profile_data, staff_required_data = [], []

    for row in range(data_start_row, 100):  # 用100作为最大行数限制
        values = [sheet.cell(row=row, column=col).value for col in duty_profile_cols]
        if all(v is None for v in values):
            break
        duty_profile_data.append(dict(zip(duty_profile_headers, values)))

    for row in range(data_start_row, 100):
        values = [sheet.cell(row=row, column=col).value for col in staff_required_cols]
        if all(v is None for v in values):
            break
        staff_required_data.append(dict(zip(staff_required_headers, values)))

    all_json_results["[Duty Profile]"] = duty_profile_data
    all_json_results["[Staff Required]"] = staff_required_data

    # === DutyLeaveRequest ===
    sheet = workbook["DutyLeaveRequest"]
    name_col = 2
    start_col = 3
    header_row = 4
    data_start_row = 5
    data_end_row = 24
    end_col = sheet.max_column

    date_headers = []
    for col in range(start_col, end_col + 1):
        val = sheet.cell(row=header_row, column=col).value
        if isinstance(val, datetime):
            date_headers.append(val.strftime('%Y-%m-%d'))
        else:
            try:
                parsed = datetime.strptime(val.strip(), "%d/%m")
                date_headers.append(f"2024-{parsed.month:02d}-{parsed.day:02d}")
            except:
                date_headers.append(str(val).strip())

    leave_request_data = []
    for row in range(data_start_row, data_end_row + 1):
        name = sheet.cell(row=row, column=name_col).value
        if not name:
            continue
        schedule = {}
        for i, col in enumerate(range(start_col, start_col + len(date_headers))):
            val = sheet.cell(row=row, column=col).value
            if val is not None:
                schedule[date_headers[i]] = val
        leave_request_data.append({"Name": name, "Schedule": schedule})
    all_json_results["[Duty/Leave Request]"] = leave_request_data

    # === HouseRule Sheet ===
    sheet = workbook["HouseRule"]
    section_titles = [
        "[Special Duty Count]", "[Day Between]", "[Staff pairing On Same Day]",
        "[Duty among Seniority On Same Day]", "[Daily Manpower]",
        "[Consecutive Duty Max Count]", "[Conditional Duty On Same Day]",
        "[DutyPattern Assignment]", "[Team pairing]"
    ]
    current_section, data = None, []

    def store_house_rule(section_name, rows):
        if not rows or len(rows) < 2:
            return []
        headers = [str(c).strip() if c else f"Unnamed:{i}" for i, c in enumerate(rows[0])]
        section_data = []
        for row in rows[1:]:
            if all(cell is None for cell in row):
                continue
            record = {}
            for i in range(len(headers)):
                val = row[i] if i < len(row) else None
                if isinstance(val, datetime):
                    val = val.strftime('%Y-%m-%d %H:%M:%S')
                record[headers[i]] = val
            section_data.append(record)
        return section_data

    for row in sheet.iter_rows(values_only=True):
        if all(cell is None for cell in row):
            continue
        first_cell = str(row[0]).strip() if row[0] else ""
        if first_cell in section_titles:
            if current_section and data:
                all_json_results[current_section] = store_house_rule(current_section, data)
            current_section = first_cell
            data = []
        elif current_section:
            data.append(row)
    if current_section and data:
        all_json_results[current_section] = store_house_rule(current_section, data)

    # === 调用 OpenRouter 分析内容 ===
    output_string = ""
    for section, content in all_json_results.items():
        json_str = json.dumps(content, indent=2, ensure_ascii=False)
        if len(json_str) > MAX_PROMPT_LENGTH:
            output_string += f"\n--- {section} 内容过长，已跳过分析（字符数={len(json_str)}） ---\n"
            continue
        prompt = f"以下是表格 {section} 的 JSON 数据：\n{json_str}\n请逐项详细描述该表格的全部信息，确保不遗漏任何日期、人员、值班等内容，并以自然、易懂的中文表达。"
        try:
            explanation = query_openrouter(prompt)
            output_string += f"\n=== {section} 的中文描述 ===\n{explanation}\n"
        except Exception as e:
            output_string += f"\n--- {section} 分析失败: {str(e)}\n"

    return output_string

